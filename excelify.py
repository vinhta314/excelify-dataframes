import pandas as pd

from openpyxl.styles import PatternFill, Color


class Pointer:
    def __init__(self, x, y):
        self.x = x
        self.y = y
        self._x0 = None
        self._y0 = None
        self._memorise = True

    def __repr__(self):
        return f'({self.x}, {self.y})'

    def __eq__(self, other):
        return self.x == other.x and self.y == other.y

    def move(self, distance=1, axis=0, committed=False):
        if self._memorise:
            self._memorise_position()
            self._move_axis(distance, axis)
            self._memorise = False
        else:
            self._move_axis(distance, axis)

        if committed:
            self.commit()

    def commit(self):
        if not self._memorise:
            self._wipe_memory()
            self._memorise = True

    def rollback(self, axis=None):
        if not self._memorise:
            self._restore_position(axis)
            self._wipe_memory()
            self._memorise = True

    def clone(self, distance=0, axis=0):
        p_clone = Pointer(self.x, self.y)
        p_clone.move(distance, axis, committed=True)
        return p_clone

    def _move_axis(self, distance=1, axis=0):
        if axis == 0:
            self.x += distance
        elif axis == 1:
            self.y += distance

    def _memorise_position(self):
        self._x0 = self.x
        self._y0 = self.y

    def _restore_position(self, axis):
        if axis != 1:
            self.x = self._x0
        elif axis != 0:
            self.y = self._y0

    def _wipe_memory(self):
        self._x0 = None
        self._y0 = None


class ExcelifyDataFrames:
    def __init__(self, wb, title, position):
        self.title = title
        self.position = position
        self.sheet = wb.create_sheet(title, position)

    def add_dataframe(self, dataframe, x, y, header=True, index=False):
        pointer = Pointer(x, y)

        if header:
            self._process_index(dataframe.columns, pointer, axis=0)

        if index:
            self._process_index(dataframe.index, pointer, axis=1)

        self._add_dataframe_data(dataframe, pointer)

    def _process_index(self, index, pointer, axis):
        if isinstance(index, pd.MultiIndex):
            self._add_higher_level_indexes(index, pointer, axis)
            root_index = self._get_root_index_from_multilevel(index)
            self._add_root_index(root_index, pointer, axis)
        else:
            self._add_root_index(index, pointer, axis=axis)

    def _add_higher_level_indexes(self, index, pointer, axis):
        for level, label in zip(index.levels[:-1], index.labels[:-1]):
            self._add_and_merge_higher_index(level, label, pointer, axis)

    def _add_root_index(self, values, pointer, axis=0):
        for i, value in enumerate(values):
            self.sheet.cell(column=pointer.x, row=pointer.y, value=value)
            pointer.move(axis=axis)

        pointer.rollback()
        pointer.move(axis=axis, commited=True)

    def _add_dataframe_data(self, dataframe, pointer):
        for j, row in dataframe.iterrows():
            for i, cell_value in enumerate(row):
                self.sheet.cell(column=pointer.x+i, row=pointer.y+j, value=cell_value)

    def apply_border(self, cell_edges, border_style):
        for cell_edge in cell_edges:
            x = cell_edge["position"][0]
            y = cell_edge["position"][1]
            style = border_style[cell_edge["borders"]]

            self.sheet.cell(row=y, column=x).border = style

    def fill_colour(self, cell_array, rgb):
        colour_fill = PatternFill(fgColor=Color(rgb))

        for cell in cell_array:
            self.sheet.cell(row=cell[1], column=cell[0]).fill = colour_fill

    def _add_and_merge_higher_index(self, values, keys, pointer, axis):
        hold = pointer.clone()

        for i, key in enumerate(keys):
            if i == len(keys) - 1 or key != keys[key + 1]:
                pointer.move(i, axis)
                self.sheet.cell(column=pointer.x, row=pointer.y, value=values[key])

                if hold != pointer:
                    self.sheet.merge_cells(
                        start_row=hold.y, start_column=hold.x,
                        end_row=pointer.y, end_column=pointer.x
                    )

                hold = pointer.clone(i + 1, axis)

        pointer.rollback()
        pointer.move(1, axis, commited=True)

    @staticmethod
    def _get_root_index_from_multilevel(multi_index):
        root_level = multi_index.levels[-1]
        root_label = multi_index.labels[-1]

        return [root_level[index] for index in root_label]
