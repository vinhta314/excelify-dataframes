from openpyxl.styles import PatternFill, Color


class ExcelSheet:
    def __init__(self, wb, title, position):
        self.title = title
        self.position = position
        self.sheet = wb.create_sheet(title, position)

    def add_dataframe_values(self, dataframe, x0, y0):
        for j, row in dataframe.iterrows():
            for i, cell_value in enumerate(row):
                self.sheet.cell(column=x0+i, row=y0-j, value=cell_value)

    def apply_border(self, cell_edges, border_style):
        for cell_edge in cell_edges:
            x = cell_edge["position"][0]
            y = cell_edge["position"][1]
            style = border_style[cell_edge["borders"]]

            self.sheet.cell(row=y, column=x).border = style

    def fill_colour(self, cell_array, rgb):
        colour_fill = PatternFill(fgColor=Color(rgb))

        for cell in cell_array:
            self.sheet.cell(row=cell[1], column=cell[0]).fill = colour_fill
